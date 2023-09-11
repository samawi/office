import streamlit as st
import pandas as pd
import plotly.express as px
import altair as alt
import numpy as np
from st_aggrid import GridOptionsBuilder, AgGrid, GridUpdateMode, DataReturnMode

import openpyxl
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from itertools import islice

import xlwings as xw

# set input file names
input_filename_scenario_01 = 'tenancy_list_scenario_01C.xlsx'
input_filename_scenario_02 = 'tenancy_list_scenario_02b.xlsx'

#set output file name
output_filename = 'Revenue 2024 2033 simulation yearly.xlsx'

# set input file path
input_path = '/Users/mawan/OneDrive/Office/AMG/Development/office/revenue-streamlit/'

# open excel files
wb_scenario_01 = openpyxl.load_workbook(input_path + input_filename_scenario_01, data_only=True)
wb_scenario_02 = openpyxl.load_workbook(input_path + input_filename_scenario_02, data_only=True)

# get all sheet names
sheet_names_scenario_01 = wb_scenario_01.sheetnames
sheet_names_scenario_02 = wb_scenario_02.sheetnames

# assign variables to 'params" sheet
ws_params_scenario_01 = wb_scenario_01['params']
ws_params_scenario_02 = wb_scenario_02['params']

# assign variables to 'tenant' sheet
ws_tenant_scenario_01 = wb_scenario_01['tenant']
ws_tenant_scenario_02 = wb_scenario_02['tenant']

# read parameters at cell C4 as date_start
date_start_scenario_01 = ws_params_scenario_01['C4'].value
date_start_scenario_02 = ws_params_scenario_02['C4'].value

# read parameter at cell C5 as date_end
date_end_scenario_01 = ws_params_scenario_01['C5'].value
date_end_scenario_02 = ws_params_scenario_02['C5'].value

# read parameter at cell C6 as area_rentable_office
area_rentable_office_scenario_01 = ws_params_scenario_01['C6'].value
area_rentable_office_scenario_02 = ws_params_scenario_02['C6'].value

# define a function to read a sheet into a dataframe
def read_sheet_to_df(wb, sheet_name):
    df = pd.DataFrame(wb[sheet_name].values)
    df.columns = df.iloc[0]
    df = df.iloc[1:]
    return df

# read content of 'tenant' sheet into a dataframe
df_tenant_scenario_01 = read_sheet_to_df(wb_scenario_01, 'tenant')
df_tenant_scenario_02 = read_sheet_to_df(wb_scenario_02, 'tenant')

# generate daterange called 'reporting_date_range' based on date start and date end
reporting_date_range_scenario_01 = pd.date_range(date_start_scenario_01, date_end_scenario_01, freq='MS')

# create empty dataframe called 'df_report_rental_charge' using datetime index for the daterange period
df_report_rental_charge_scenario_01 = pd.DataFrame(index=reporting_date_range_scenario_01)

# create empty dataframe called 'df_report_sc_charge' using datetime index for the daterange period
df_report_sc_charge_scenario_01 = pd.DataFrame(index=reporting_date_range_scenario_01)

# create empty dataframe called 'df_report_sc_charge_rate' using datetime index for the daterange period
df_report_sc_charge_rate_scenario_01 = pd.DataFrame(index=reporting_date_range_scenario_01)

# create empty dataframe called 'df_report_occupancy' using datetime index for the daterange period
df_report_occupancy_scenario_01 = pd.DataFrame(index=reporting_date_range_scenario_01)

# set 'date' as the index of the dataframe
df_report_rental_charge_scenario_01.index.name = 'date'
df_report_sc_charge_scenario_01.index.name = 'date'
df_report_sc_charge_rate_scenario_01.index.name = 'date'
df_report_occupancy_scenario_01.index.name = 'date'

sc_data = [
        [datetime(2018, 4, 1), 65000.00],
        [datetime(2020, 4, 1), 70000.00],
        [datetime(2022, 4, 1), 70000.00],
        [datetime(2024, 4, 1), 75000.00],
        [datetime(2027, 4, 1), 80000.00],
        [datetime(2030, 4, 1), 85000.00],
        [datetime(2033, 4, 1), 90000.00],
]

product_type = 'office'

# select rows according to product_type
df_tenant_scenario_01 = df_tenant_scenario_01[df_tenant_scenario_01['Product_Type'] == product_type]

substring = '-'

iteration = 0

# iterate each row in df_tenant
for index, row in df_tenant_scenario_02.iterrows():

    # print iteration to console
    iteration += 1
    print('done')


    # define initial value of variables
    vacant = False

    # read columns 'Area', 'Rental_Rate', 'SC_Rate', 'Tenant', 'Chg_Type', 'Start', 'End'
    area = row['Area']
    rental_rate = row['Rental_Rate']
    sc_rate = row['SC_Rate']
    tenant = row['Tenant']
    chg_type = row['Chg_Type']
    start = row['Start']
    end = row['End']

    # if area is None or substring then set area to 0
    if area is None or substring in str(area):
        area = 0.0

    # if rental_rate is None or substring then set rental_rate to 0
    if rental_rate is None or substring in str(rental_rate):
        rental_rate = 0.0

    # if sc_rate is None or substring then set sc_rate to 0
    if sc_rate is None or substring in str(sc_rate):
        sc_rate = 0.0

    # if chg_type is 'A', then calculate rental charge
    if chg_type == 'A':
        rental_charge = area * rental_rate
    else:
        rental_charge = rental_rate

    # calculate service charge
    sc_charge = area * sc_rate

    # if end is None or substring then set vacant to True and end to date_end
    if end is None or substring in str(end):
        vacant = True
        end = date_end_scenario_02

print('Done')
