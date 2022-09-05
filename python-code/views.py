from django.shortcuts import render
import openpyxl


def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:

        # read excel file

        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        #print(sheets)

        # getting a particular sheet
        param_sht = wb["params"]
        sht1 = wb["tenant"]
        rental_sht = wb["rental"]
        yearly_rental_sheet = wb["yearly_rental"]
        sc_sht = wb["sc"]
        yearly_sc_sht = wb["yearly_sc"]
        total_sheet = wb["total_rev"]
        occ_sheet = wb["occ_rate"]

        #print(param_sht)

        # getting active sheet
        #active_sheet = wb.active
        #print(active_sheet)

        # reading parameters
        start_date = param_sht["C4"].value
        end_date = param_sht['C5'].value
        rentable_office_area = param_sht["C6"].value
        sum_report = param_sht["C7"].value

        # set excel column name for values
        level_addr_col = 'B'        # Floor
        zone_addr_col = 'C'         # Zone
        area_addr_col = 'E'         # Area
        rental_rate_addr_col = 'H'  # Rental rate
        sc_rate_addr_col = 'I'      # SC rate
        lcd_addr_col = 'F'          # LCD
        led_addr_col = 'G'          # LED
        start_show_col = 'M'

        print(f"Start Date: {start_date}")
        print(f"End Date: {end_date}")
        print(f"rentable Office Area: {rentable_office_area}")
        print(f"Sum Report: {sum_report}")

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in param_sht.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'myapp/index.html', {"excel_data":excel_data})









